import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side


def example():
    book=openpyxl.load_workbook("test1.xlsx")
    sheet=book.worksheets[0]
    
    header_font=Font(bold=True, size=12, color="FFFFFF")
    header_fill=PatternFill("darkGray", fgColor="4F8180")
    header_aligment=Alignment(horizontal="center", vertical="center")
    thin=Side(style="thin", color="555555")
    boarder=Border(left=thin, right=thin, top=thin, bottom=thin)
    
    sheet.row_dimensions[1].height=24
    
    for cell in sheet[1]:
        cell.font=header_font
        cell.alignment=header_aligment
        cell.fill=header_fill
        cell.border=boarder
    
    sheet.column_dimensions["A"].width=40
    sheet.column_dimensions["B"].width=50
    sheet.column_dimensions["C"].width=30
    sheet.column_dimensions["D"].width=10
    
    book.save("test_styled.xlsx")