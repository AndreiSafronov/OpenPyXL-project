import openpyxl
from examples.data_faker import data_samples


def example():
    book = openpyxl.Workbook()
    book.remove(book.active)
    book.create_sheet("Персонал")
    book.create_sheet("Студенты")
    book.create_sheet("Преподаватели", 0)

    for worksheet in book.worksheets:
        worksheet.append(["ФИО", "Адрес", "Город", "Статус"])

    for sheet in book.worksheets:
        for row in data_samples():
            sheet.append(row)

    book.save("test.xlsx")
